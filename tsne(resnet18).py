import torch.optim as optim
import torch
import torch.nn as nn
import torch.nn.parallel
import torch.optim
import torch.utils.data
import torch.utils.data.distributed
import torchvision.transforms as transforms
import torchvision.datasets as datasets
import torchvision.models
import torchvision.io as io
from torch.autograd import Variable
import os
import shutil
import datetime
import dispatch
import numpy
import sys
from sklearn import preprocessing
from sklearn.manifold import TSNE
import operator
import torch
import torchvision
import matplotlib.pyplot as plt
from sklearn.metrics import f1_score, recall_score, precision_score, accuracy_score
import seaborn
import openpyxl
import pandas
import mpl_toolkits

# 设置超参数
os.environ['PYTORCH_CUDA_ALLOC_CONF'] = 'max_split_size_mb:4096'  # 防止gpu記憶體碎片化
FOLDER = str('./experiment'+str(datetime.datetime.now()).replace(':', '-'))
SOURCEFOLDER = str('./picsrcdispatched')
BATCH_SIZE = 16
EPOCHS =10
DEVICE = torch.device('cuda' if torch.cuda.is_available() else 'cpu')
# TRAINRATIO = 4
# VALIDATERATIO =  4
# TESTRATIO =  2
CLASSES = ('Aspergillus flavus', 'Aspergillus fumigatus',
           'Aspergillus niger', 'Aspergillus terreus')
CLASSES1 = ('Aspergillus\nflavus', 'Aspergillus\nfumigatus',
            'Aspergillus\nniger', 'Aspergillus\nterreus')
CLASSES2 = ('A. flavus', 'A. fumigatus',
            'A. niger', 'A. terreus')

# 定義結果類別


class PictureResult:
    def __init__(self, label, species, utility):
        self.label = label
        self.species = species
        self.utility = utility

    def alltest(self, flavus, fumigatus, niger, terreus, predict):
        self.flavus = flavus
        self.fumigatus = fumigatus
        self.niger = niger
        self.terreus = terreus
        self.predict = predict


# 宣告結果陣列,預測陣列,結果哈希表,結果物件陣列
actual = []
predict = []
actual_number = []
predict_number = []
species = []
utility = []
result_hash = dict()
a = []

# 準備結果報告excel檔案

report = openpyxl.Workbook()
sheet = report.worksheets[0]
value=report.create_sheet("value")
sheet.cell(row=1, column=1, value='Image File Name')
for i in range(4):
    sheet.cell(row=1, column=i+4, value=CLASSES[i]+" percentage")
sheet.cell(row=1, column=2, value='Species')
sheet.cell(row=1, column=3, value='Utility')
sheet.cell(row=1, column=8, value='predict')
sheet.cell(row=1, column=9, value='label')
sheet.cell(row=1, column=10, value='Ambiguity')
sheet.freeze_panes = 'B2'
value.cell(row=1, column=2, value='Accuracy')
value.cell(row=1, column=3, value='Precision')
value.cell(row=1, column=4, value='Recall')
value.cell(row=1, column=5, value='F1-score')

value.cell(row=2, column=1, value=CLASSES[0])
value.cell(row=3, column=1, value=CLASSES[1])
value.cell(row=4, column=1, value=CLASSES[2])
value.cell(row=5, column=1, value=CLASSES[3])

# confusion_excel =report.worksheets[1]

# 準備loss,accuracy對iteration儲存空間
train_loss = numpy.empty([176*EPOCHS])
train_loss_avg = numpy.empty([176*EPOCHS])
train_accuracy = numpy.empty([176*EPOCHS])
validate_loss = numpy.empty([176*EPOCHS])
validate_loss_avg = numpy.empty([176*EPOCHS])
validate_accuracy = numpy.empty([176*EPOCHS])
# 建立資料夾
os.mkdir(FOLDER)
os.mkdir(FOLDER+'/missed')
os.mkdir(FOLDER+'/missed/Aspergillus flavus')
os.mkdir(FOLDER+'/missed/Aspergillus fumigatus')
os.mkdir(FOLDER+'/missed/Aspergillus niger')
os.mkdir(FOLDER+'/missed/Aspergillus terreus')


# 從./picsrc處讀入照片
# dispatch.dispatchbyratio(TRAINRATIO,VALIDATERATIO,TESTRATIO,FOLDER)


# 数据预处理

transform = transforms.Compose([
    transforms.Resize((128, 128)),
    # transforms.RandomVerticalFlip(),
    # transforms.RandomCrop(50),
    # transforms.ColorJitter(brightness=0.5, contrast=0.5, hue=0.5),
    transforms.ToTensor(),
    transforms.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])

])
transform_test = transforms.Compose([
    transforms.Resize((128, 128)),
    transforms.ToTensor(),
    transforms.Normalize([0.5, 0.5, 0.5], [0.5, 0.5, 0.5])
])
# 读取数据
dataset_train = datasets.ImageFolder(SOURCEFOLDER+'/train', transform)
# print(dataset_train.imgs)
# 对应文件夹的label
# print(dataset_train.class_to_idx)
dataset_test = datasets.ImageFolder(SOURCEFOLDER+'/validate', transform_test)
# 对应文件夹的label
# print(dataset_test.class_to_idx)
dataset_trial = datasets.ImageFolder(SOURCEFOLDER+'/test', transform_test)
dataset_alltrial = datasets.ImageFolder(
    SOURCEFOLDER+'/alltest', transform_test)


# 导入数据
train_loader = torch.utils.data.DataLoader(
    dataset_train, batch_size=BATCH_SIZE, shuffle=True)
test_loader = torch.utils.data.DataLoader(
    dataset_test, batch_size=BATCH_SIZE, shuffle=False)
trial_loader = torch.utils.data.DataLoader(
    dataset_trial, batch_size=BATCH_SIZE, shuffle=False)
alltrial_loader = torch.utils.data.DataLoader(
    dataset_alltrial, batch_size=BATCH_SIZE, shuffle=False)

train_plot_loader = torch.utils.data.DataLoader(
    dataset_train, batch_size=1, shuffle=False)
test_plot_loader = torch.utils.data.DataLoader(
    dataset_test, batch_size=1, shuffle=False)
trial_plot_loader = torch.utils.data.DataLoader(
    dataset_trial, batch_size=1, shuffle=False)
alltrial_plot_loader = torch.utils.data.DataLoader(
    dataset_alltrial, batch_size=1, shuffle=False)


# 取出dataset的(資料image,標籤targets)
x_image=[]
y_label=[]
with torch.no_grad():
    for i, (picture, label) in enumerate(trial_plot_loader):
        print(i)
        x_image.append(picture)
        
        #picture, label = Variable(picture), Variable(label)
        if i == 0:
            #x_image_original = picture
            y_label_original = label
        else:
            #x_image_original = torch.cat((x_image_original, picture), 0)
            y_label_original = torch.cat((y_label_original, label), 0)
    
    x_image_original=torch.stack(x_image,dim=0)
 
# 資料image,標籤targets降維
    x_image_original = x_image_original.view(len(x_image_original), -1)
    y_label_original = y_label_original.view(len(y_label_original), -1)

# 設定學習率
modellr = 1e-3
# 準備訓練後資料儲存空間
projected_picture_in_r4 = torch.zeros(1, 4)


# 選擇損失函數
criterion = nn.CrossEntropyLoss()

# 实例化類神經網路并且移动到GPU
model = torchvision.models.resnet18(pretrained=False)  # resnet18
num_ftrs = model.fc.in_features
# 本次任務需要辨認4種真菌
model.fc = nn.Linear(num_ftrs, 4)
model.to(DEVICE)

# 选择Adam优化器，学习率
optimizer = optim.AdamW(model.parameters(), lr=modellr, weight_decay=0.01)

# 取出dataset的(資料image,標籤targets)





def adjust_learning_rate(optimizer, epoch):  # 將學習率逐步調降以避免loss spike情形發生
    """Sets the learning rate to the initial LR decayed """
    lr_list = [1e-3, 1e-4, 1e-4, 1e-5, 1e-5, 1e-5, 5e-7, 1e-7]
    for param_group in optimizer.param_groups:
        if epoch < len(lr_list)+1:
            param_group['lr'] = lr_list[epoch-1]
        else:
            param_group['lr'] = lr_list[-1]


# 定义训练过程
def train(model, device, train_loader, optimizer, epoch):
    global train_loss, train_accuracy
    model.train()
    sum_loss = 0
    correct = 0
    acc = 0
    i = 0
    num = 0
    total_num = len(train_loader.dataset)
    print(total_num, len(train_loader))
    for batch_idx, (picture, label) in enumerate(train_loader):
        if epoch == 1:
            for j in range(0, 16):
                if j+16*batch_idx >= len(train_loader.dataset.imgs):
                    break
                a.append(PictureResult(
                    train_loader.dataset.imgs[j+16*batch_idx][1], CLASSES[train_loader.dataset.imgs[j+16*batch_idx][1]], 'Train'))
                result_hash[train_loader.dataset.imgs[j +
                                                      16*batch_idx][0].replace('train','alltest')] = a[j+16*batch_idx]
        picture, label = Variable(picture).to(
            device), Variable(label).to(device)
        num += picture.shape[0]
        output = model(picture)
        loss = criterion(output, label)
        _, pred = torch.max(output.data, 1)
        correct += torch.sum(pred == label)
        train_accuracy[i+(epoch-1)*176] = correct/num
        optimizer.zero_grad()  # 清除暫存參數
        loss.backward()  # 求梯度
        optimizer.step()  # 梯度反向傳播,更新類神經網路參數
        print_loss = loss.data.item()
        sum_loss += print_loss
        if (batch_idx + 1) % 50 == 0:
            print('Train Epoch: {} [{}/{} ({:.0f}%)]\tLoss: {:.6f}'.format(
                epoch, (batch_idx + 1) *
                len(picture), len(train_loader.dataset),
                100. * (batch_idx + 1) / len(train_loader), loss.item()))
        train_loss[i+(epoch-1)*176] = sum_loss/(i+1)
        i += 1
        del picture, label
    ave_loss = sum_loss / len(train_loader)
    torch.cuda.empty_cache()
    print('epoch:{},loss:{}'.format(epoch, ave_loss))

# 定义驗證过程


def val(model, device, val_loader, epoch):
    global validate_loss, validate_accuracy
    model.eval()
    test_loss = 0
    correct = 0
    acc = 0
    num = 0
    i = 0
    total_num = len(val_loader.dataset)
    print(total_num, len(val_loader))
    with torch.no_grad():
        for batch_idx, (data, target) in enumerate(val_loader):
            if epoch == 1:
                for j in range(0, 16):
                    if j+16*batch_idx >= len(val_loader.dataset.imgs):
                        break
                    a.append(PictureResult(
                        val_loader.dataset.imgs[j+16*batch_idx][1], CLASSES[val_loader.dataset.imgs[j+16*batch_idx][1]], 'Validate'))
                    result_hash[val_loader.dataset.imgs[j+16 *
                                                         batch_idx][0].replace('validate','alltest')] = a[2813+j+16*batch_idx]
            data, target = Variable(data).to(
                device), Variable(target).to(device)
            num += data.shape[0]
            output = model(data)
            loss = criterion(output, target)
            _, pred = torch.max(output.data, 1)
            correct += torch.sum(pred == target)
            validate_accuracy[i+(epoch-1)*176] = correct/num
            print_loss = loss.data.item()
            test_loss += print_loss
            validate_loss[i+(epoch-1)*176] = test_loss/(i+1)
            i += 1
            del data, target  # 釋放記憶體
       # correct = correct.data.item()
        acc = correct / total_num
        avgloss = test_loss / len(val_loader)
        print('\n驗證集: Average loss: {:.4f}, Accuracy: {}/{} ({:.0f}%)\n'.format(
            avgloss, correct, len(val_loader.dataset), 100 * acc))
    return 100 * (acc.data.item()),correct,total_num

# 定義測試過程(用cpu避免memory不足)


def test(model, device, test_loader):
    with torch.no_grad():
        model = torch.load(FOLDER+'/model.pth')
        model.eval()
        # model.to(device)
        print(len(test_loader))  # 对应文件夹的label
        count = 0
        right = 0
        for index in range(len(test_loader)):
            a.append(PictureResult(
                test_loader.imgs[index][1], CLASSES[test_loader.imgs[index][1]], 'Test'))
            result_hash[test_loader.imgs[index][0].replace('test','alltest')] = a[5627+index]
            item = test_loader[index]
            img, label = item
            img.unsqueeze_(0)
            data = Variable(img).to(device)
            output = model(data)
            if index == 0:
                projected_picture = output
            else:
                projected_picture = torch.cat((projected_picture, output), 0)

            _, pred = torch.max(output.data, 1)
            print('Image Name:{},predict:{}'.format(
                test_loader.imgs[index][0], CLASSES[pred.data.item()]))
    #       if(test_loader.imgs[index][0].find(CLASSES[pred.data.item()])!=-1):
            actual.append(CLASSES2[label])
            actual_number.append(label)
            predict.append(CLASSES1[pred.cpu().item()])
            predict_number.append(pred.cpu().item())
            if (label == pred):
                count += 1
            index += 1
            del data
    print('已比對圖片數目:{},正確圖片數目:{}'.format(index, count))

    return index, count, projected_picture


# 定義所有集合測試過程(用cpu避免memory不足)
def alltest(model, device, test_loader):
    with torch.no_grad():
        model = torch.load(FOLDER+'/model.pth')
        model.eval()
        # model.to(device)
        print(len(test_loader))
        count = 0
        right = 0
        for index in range(len(test_loader)):
            item = test_loader[index]
            img, label = item
            # torch.nn只支持小批量处理 (mini-batches）。整个 torch.nn 包只支持小批量样本的输入，不支持单个样本的输入。比如，nn.Conv2d 接受一个4维的张量，即nSamples x nChannels x Height x Width，如果是一个单独的样本，只需要使用input.unsqueeze(0) 来添加一个“假的”批大小维度
            img.unsqueeze_(0)
            data = Variable(img).to(device)
            output = model(data)
            #if index == 0:
            #    projected_picture = output
            #else:
            #    projected_picture = torch.cat((projected_picture, output), 0)
            _, pred = torch.max(output.data, 1)
            print('Image Name:{},predict:{}'.format(
                test_loader.imgs[index][0], CLASSES[pred.data.item()]))
            
            
            normedoutput = torch.nn.functional.softmax(output.squeeze(), 0)
            result_hash[test_loader.imgs[index][0]].flavus=normedoutput[0].cpu().item()
            result_hash[test_loader.imgs[index][0]].fumigatus=normedoutput[1].cpu().item()
            result_hash[test_loader.imgs[index][0]].niger=normedoutput[2].cpu().item()
            result_hash[test_loader.imgs[index][0]].terreus=normedoutput[3].cpu().item()
            result_hash[test_loader.imgs[index][0]].predict=pred.cpu().item()

            # normedoutput.cpu()
            # normedoutputnumpy = normedoutput.numpy()
            '''
            sheet.cell(row=index+2, column=1, value=test_loader.imgs[index][0])
            sheet.cell(row=index+2, column=2,
                       value=normedoutput[0].cpu().item())
            sheet.cell(row=index+2, column=3,
                       value=normedoutput[1].cpu().item())
            sheet.cell(row=index+2, column=4,
                       value=normedoutput[2].cpu().item())
            sheet.cell(row=index+2, column=5,
                       value=normedoutput[3].cpu().item())
            '''
            if (label == pred):
                count += 1
            else:
                shutil.copy(
                    test_loader.imgs[index][0], FOLDER+'/missed/'+CLASSES[pred.data.item()])
                
            index += 1
            del data
#    print('已比對圖片數目:{},正確圖片數目:{}'.format(index,count))
    return index, count

# 填入excel圖表
def excel(loader):
    with torch.no_grad():
        for index in range(len(loader)):
            sheet.cell(row=index+2, column=1, value=loader.imgs[index][0]).alignment=openpyxl.styles.Alignment(horizontal='right')
            sheet.cell(row=index+2, column=2, value=result_hash[loader.imgs[index][0]].species)
            sheet.cell(row=index+2, column=3, value=result_hash[loader.imgs[index][0]].utility)
            sheet.cell(row=index+2, column=4, value=result_hash[loader.imgs[index][0]].flavus)
            sheet.cell(row=index+2, column=5, value=result_hash[loader.imgs[index][0]].fumigatus)
            sheet.cell(row=index+2, column=6, value=result_hash[loader.imgs[index][0]].niger)
            sheet.cell(row=index+2, column=7, value=result_hash[loader.imgs[index][0]].terreus)
            sheet.cell(row=index+2, column=8, value=result_hash[loader.imgs[index][0]].predict)
            sheet.cell(row=index+2, column=9, value=result_hash[loader.imgs[index][0]].label)
            if  result_hash[loader.imgs[index][0]].predict!=result_hash[loader.imgs[index][0]].label:
                A=[result_hash[loader.imgs[index][0]].flavus,result_hash[loader.imgs[index][0]].fumigatus,result_hash[loader.imgs[index][0]].niger,result_hash[loader.imgs[index][0]].terreus]
                A.sort()
                sheet.cell(row=index+2, column=10,value=A[2]/A[3])
                Yellow = ['ffeb9c']
                Ink = openpyxl.styles.PatternFill('solid', fgColor=Yellow[0])
                sheet.cell(row=index+2, column=1).fill = Ink
            if  result_hash[loader.imgs[index][0]].utility=='Train':
                Blue = ['00FFFF']
                Ink = openpyxl.styles.PatternFill('solid', fgColor=Blue[0])
                sheet.cell(row=index+2, column=3).fill = Ink
            if  result_hash[loader.imgs[index][0]].utility=='Validate':
                Green = ['FFAF60']
                Ink = openpyxl.styles.PatternFill('solid', fgColor=Green[0])
                sheet.cell(row=index+2, column=3).fill = Ink
            if  result_hash[loader.imgs[index][0]].utility=='Test':
                Red = ['FFBFFF']
                Ink = openpyxl.styles.PatternFill('solid', fgColor=Red[0])
                sheet.cell(row=index+2, column=3).fill = Ink







# 训练
# accuracy = 0
finalepoch=0
for epoch in range(1, EPOCHS + 1):
    adjust_learning_rate(optimizer, epoch)
    train(model, DEVICE, train_loader, optimizer, epoch)
    _,a1,b=val(model, DEVICE, test_loader, epoch)
    if epoch>5 and (b-a1.item())<5:
        train_accuracy= train_accuracy[:176*epoch]
        train_loss=train_loss[:176*epoch]
        validate_accuracy=validate_accuracy[:176*epoch]
        validate_loss=validate_loss[:176*epoch]
        finalepoch=epoch  
        break
    else:
        finalepoch=EPOCHS      
# while accuracy <= 90:
#    train(model, DEVICE, train_loader, optimizer, epoch)
#    accuracy=val(model, DEVICE, test_loader)
torch.save(model, FOLDER+'/model.pth')

# 測試集測試
index1, count1 , projected_picture_in_r4= test(model, DEVICE, dataset_trial)

# 所有集合測試,匯出錯誤圖片
index2, count2 = alltest(
    model, DEVICE, dataset_alltrial)
excel(dataset_alltrial)
# 產生報告
# print('訓練集:驗證集:測試集={}:{}:{}'.format(TRAINRATIO,VALIDATERATIO,TESTRATIO))
print('已比對測試圖片數目:{},正確圖片數目:{}'.format(index1, count1))
print('測試集正確率:{}%'.format(count1/index1*100))
print('已比對所有圖片數目:{},正確圖片數目:{}'.format(index2, count2))
print('所有圖片正確率:{}%'.format(count2/index2*100))
#設定mathplot圖表字體,大小
plt.rcParams['font.family']='Times New Roman'
plt.rcParams['font.size']='24'

# 繪製tSNE圖表

# tsne manifold embedding
tsne = TSNE(n_components=2, random_state=42)
original_tsne_plot = tsne.fit_transform(x_image_original.cpu())
trained_tsne_plot = tsne.fit_transform(projected_picture_in_r4.detach().cpu())
# rescale to -1~1
scaler = preprocessing.MinMaxScaler(feature_range=(-1, 1))
result_original = scaler.fit_transform(original_tsne_plot)
result_trained = scaler.fit_transform(trained_tsne_plot)


fig, ax = plt.subplots(2, 1, figsize=(15, 30))
ax[0].set_title('t-SNE scattered plot of raw fungal images')
ax[1].set_title(
    't-SNE scattered plot of fungal images through trained Resnet neural networks')
scatter_0 = ax[0].scatter(
    result_original[:, 0], result_original[:, 1], c=y_label_original.cpu(), s=10)
scatter_1 = ax[1].scatter(result_trained[:, 0],
                          result_trained[:, 1], c=y_label_original.cpu(), s=10)
handles, _ = scatter_0.legend_elements(prop='colors')
handles, _ = scatter_1.legend_elements(prop='colors')
ax[0].xaxis.set_ticks([])
ax[0].yaxis.set_ticks([])
ax[1].xaxis.set_ticks([])
ax[1].yaxis.set_ticks([])
ax[0].xaxis.set_ticklabels([])
ax[0].yaxis.set_ticklabels([])
ax[1].xaxis.set_ticklabels([])
ax[1].yaxis.set_ticklabels([])

ax[0].legend(handles, CLASSES)
ax[1].legend(handles, CLASSES)
fig.savefig(fname=FOLDER+'/tsne.png')

# 輸出confusion matrix
fig2, ax3 = plt.subplots(1, 1, figsize=(20, 15))
data = {'actual': actual,
        'predict': predict
        }
resultdataframe = pandas.DataFrame(data)
confusionmatrix = pandas.crosstab(resultdataframe['actual'], resultdataframe['predict'], rownames=[
                                  'Actual'], colnames=['Predict'], margins=True)
ax3=seaborn.heatmap(confusionmatrix, annot=True, ax=ax3, fmt='g',cmap='summer')
ax3.set_title('Heatmap of actual classification of fungi to\n predict of pictures of fungi through Resnet neural network')
font={'style':'italic','horizontalalignment':'center'}
plt.setp((ax3.get_xticklabels(),ax3.get_yticklabels()),**font)
fig2.savefig(fname=FOLDER+'/heatmap.png')
# 輸出f1-score
print('Accuracy:{},Precision:{},Recall:{},f1-score:{}'.format(accuracy_score(actual_number, predict_number), precision_score(actual_number,
      predict_number, average=None), recall_score(actual_number, predict_number, average=None), f1_score(actual_number, predict_number, average=None)))
x=accuracy_score(actual_number, predict_number)
for i in range(2,6):
    value.cell(row=i, column=2, value=str(accuracy_score(actual_number, predict_number)))
    value.cell(row=i, column=3, value= str(precision_score(actual_number,predict_number, average=None)[i-2]))
    value.cell(row=i, column=4, value= str(recall_score(actual_number, predict_number, average=None)[i-2]))
    value.cell(row=i, column=5, value=str(f1_score(actual_number, predict_number, average=None)[i-2]))


# 結果報告excel檔案存檔
report.save(FOLDER+'/fungal-recognition.xlsx')
# 繪製準確率對迭代次數圖表_number
fig1, ax1 = plt.subplots(2, 1, figsize=(15, 40))
ax1[0].set_title(
    'curved chart of loss to iteration during training and validating \nof Resnet neural network')
ax1[1].set_title(
    'curved chart of accuracy to iteration during training and validating \nof Resnet neural network')
curved_0 = ax1[0].plot(range(finalepoch*176), train_loss, label='train loss')
curved_1 = ax1[0].plot(range(finalepoch*176), validate_loss, label='validate loss')
twinax0 = ax1[0].twiny()
curved_2 = ax1[1].plot(range(finalepoch*176), train_accuracy,
                       label='train accuracy')
curved_3 = ax1[1].plot(
    range(finalepoch*176), validate_accuracy, label='validate accuracy')
twinax1 = ax1[1].twiny()
twinax0.set_xlim(ax1[0].get_xlim())
twinax1.set_xlim(ax1[1].get_xlim())
xtick_loc= []
for i in range(0,finalepoch+1,1):
    if i!=finalepoch:
        xtick_loc.append(i*176)
    else:
        xtick_loc.append(i*176-1)
twinax0.set_xticks(xtick_loc,range(0, finalepoch+1, 1))
twinax1.set_xticks(xtick_loc,range(0, finalepoch+1, 1))
twinax0.set_xlabel('epoch')
twinax1.set_xlabel('epoch')

ax1[0].legend()
ax1[1].legend()
ax1[0].set_xlabel('iteration')
ax1[0].set_ylabel('loss')
ax1[1].set_xlabel('iteration')
ax1[1].set_ylabel('accuracy')
fig1.savefig(fname=FOLDER+'/curved.png')
plt.show()
#print(result_hash)
# 輸出confusion
